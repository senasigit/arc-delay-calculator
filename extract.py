import openpyxl

wb = openpyxl.load_workbook('/Users/macmini/Downloads/ Arc Delay Calculator V1.25.xlsx', data_only=False)
for sheetname in wb.sheetnames:
    print(f"Sheet: {sheetname}")
    sheet = wb[sheetname]
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  {cell.coordinate}: FORMULA: {cell.value}")
                else:
                    print(f"  {cell.coordinate}: {cell.value}")
